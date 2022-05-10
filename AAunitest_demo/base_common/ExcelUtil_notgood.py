import openpyxl

class ExcelRead:
    
    def getexcel(self):
        # 获取文件
        wb = openpyxl.load_workbook('../excel/exc.xlsx')

        # 获取所有的工作表名称
        sheet = wb['tl']
#         print(sheet.max_row,sheet.max_column)
        # 通过sheet内容
        all_list = []
        for row in range(2,sheet.max_row + 1):
            temp_list = []
            for col in range(1,sheet.max_column + 1):
                temp_list.append(sheet.cell(row,col).value)
            all_list.append(temp_list)
        return all_list
    